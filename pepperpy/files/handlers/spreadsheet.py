"""Spreadsheet handler implementation"""

from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Union

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..exceptions import FileError
from ..types import FileContent, SpreadsheetStats
from .base import BaseHandler


class SpreadsheetHandler(BaseHandler):
    """Handler for CSV and Excel files"""

    async def read(self, path: Path) -> FileContent:
        """Read spreadsheet file"""
        try:
            metadata = await self._get_metadata(path)

            if path.suffix.lower() == ".csv":
                df = pd.read_csv(path)
            else:
                df = pd.read_excel(path)

            return FileContent(content=df, metadata=metadata, format="spreadsheet")
        except Exception as e:
            raise FileError(f"Failed to read spreadsheet: {str(e)}", cause=e)

    async def get_stats(self, df: pd.DataFrame) -> SpreadsheetStats:
        """Get statistical information about the data"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns

            stats = {
                "row_count": len(df),
                "column_count": len(df.columns),
                "missing_values": df.isnull().sum().to_dict(),
                "column_types": df.dtypes.to_dict(),
                "numeric_stats": df[numeric_cols].describe().to_dict(),
                "memory_usage": df.memory_usage(deep=True).sum(),
                "duplicates": df.duplicated().sum(),
            }

            return SpreadsheetStats(**stats)
        except Exception as e:
            raise FileError(f"Failed to get stats: {str(e)}", cause=e)

    async def filter_data(self, df: pd.DataFrame, conditions: Dict[str, Any]) -> pd.DataFrame:
        """Filter data based on conditions"""
        try:
            query = " and ".join(f"{k} == @conditions['{k}']" for k in conditions)
            return df.query(query)
        except Exception as e:
            raise FileError(f"Failed to filter data: {str(e)}", cause=e)

    async def group_by(
        self, df: pd.DataFrame, columns: List[str], agg_funcs: Dict[str, str]
    ) -> pd.DataFrame:
        """Group and aggregate data"""
        try:
            return df.groupby(columns).agg(agg_funcs).reset_index()
        except Exception as e:
            raise FileError(f"Failed to group data: {str(e)}", cause=e)

    async def pivot_table(
        self,
        df: pd.DataFrame,
        index: List[str],
        columns: Optional[List[str]],
        values: List[str],
        aggfunc: Literal["mean", "sum", "count", "min", "max"] = "mean",
    ) -> pd.DataFrame:
        """Create pivot table"""
        try:
            return pd.pivot_table(
                df, index=index, columns=columns, values=values, aggfunc=aggfunc
            ).reset_index()
        except Exception as e:
            raise FileError(f"Failed to create pivot table: {str(e)}", cause=e)

    async def save_excel(
        self,
        df: pd.DataFrame,
        path: Path,
        sheet_name: str = "Sheet1",
        styling: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Save data to Excel with styling"""
        try:
            writer = pd.ExcelWriter(path, engine="openpyxl")
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            if styling:
                workbook = writer.book
                worksheet = workbook[sheet_name]

                # Apply column widths
                for idx, _ in enumerate(df.columns, 1):
                    column = get_column_letter(idx)
                    worksheet.column_dimensions[column].width = styling.get("column_width", 15)

                # Apply header style
                header_fill = PatternFill(
                    start_color=styling.get("header_color", "CCCCCC"),
                    end_color=styling.get("header_color", "CCCCCC"),
                    fill_type="solid",
                )
                header_font = Font(bold=True)

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                # Apply alternating row colors
                if styling.get("alternate_rows"):
                    for row in range(2, len(df) + 2):
                        if row % 2 == 0:
                            for cell in worksheet[row]:
                                cell.fill = PatternFill(
                                    start_color=styling.get("alt_row_color", "F0F0F0"),
                                    end_color=styling.get("alt_row_color", "F0F0F0"),
                                    fill_type="solid",
                                )

            writer.close()

        except Exception as e:
            raise FileError(f"Failed to save Excel file: {str(e)}", cause=e)

    async def merge_sheets(
        self, paths: List[Path], how: str = "outer", on: Optional[Union[str, List[str]]] = None
    ) -> pd.DataFrame:
        """Merge multiple spreadsheets"""
        try:
            dfs = []
            for path in paths:
                if path.suffix.lower() == ".csv":
                    df = pd.read_csv(path)
                else:
                    df = pd.read_excel(path)
                dfs.append(df)

            result = dfs[0]
            for df in dfs[1:]:
                result = result.merge(df, how=how, on=on)

            return result
        except Exception as e:
            raise FileError(f"Failed to merge sheets: {str(e)}", cause=e)

    async def validate_data(
        self, df: pd.DataFrame, rules: Dict[str, Dict[str, Any]]
    ) -> Dict[str, List[str]]:
        """Validate data against rules"""
        errors = {}

        try:
            for column, rule in rules.items():
                if column not in df.columns:
                    continue

                column_errors = []

                # Check required
                if rule.get("required"):
                    missing = df[column].isnull().sum()
                    if missing > 0:
                        column_errors.append(f"Found {missing} missing values")

                # Check unique
                if rule.get("unique"):
                    duplicates = df[column].duplicated().sum()
                    if duplicates > 0:
                        column_errors.append(f"Found {duplicates} duplicate values")

                # Check min/max
                if "min" in rule:
                    invalid = (df[column] < rule["min"]).sum()
                    if invalid > 0:
                        column_errors.append(f"Found {invalid} values below minimum {rule['min']}")

                if "max" in rule:
                    invalid = (df[column] > rule["max"]).sum()
                    if invalid > 0:
                        column_errors.append(f"Found {invalid} values above maximum {rule['max']}")

                # Check regex pattern
                if "pattern" in rule:
                    invalid = (~df[column].str.match(rule["pattern"])).sum()
                    if invalid > 0:
                        column_errors.append(f"Found {invalid} values not matching pattern")

                if column_errors:
                    errors[column] = column_errors

            return errors

        except Exception as e:
            raise FileError(f"Failed to validate data: {str(e)}", cause=e)

    async def transform_data(
        self, df: pd.DataFrame, transformations: Dict[str, Dict[str, Any]]
    ) -> pd.DataFrame:
        """Apply transformations to data"""
        try:
            result = df.copy()

            for column, transform in transformations.items():
                if column not in result.columns:
                    continue

                # Apply type conversion
                if "type" in transform:
                    result[column] = result[column].astype(transform["type"])

                # Apply string operations
                if "case" in transform:
                    if transform["case"] == "upper":
                        result[column] = result[column].str.upper()
                    elif transform["case"] == "lower":
                        result[column] = result[column].str.lower()
                    elif transform["case"] == "title":
                        result[column] = result[column].str.title()

                # Apply numeric operations
                if "round" in transform:
                    result[column] = result[column].round(transform["round"])

                if "fill" in transform:
                    result[column] = result[column].fillna(transform["fill"])

                # Apply custom function
                if "function" in transform:
                    result[column] = result[column].apply(transform["function"])

            return result

        except Exception as e:
            raise FileError(f"Failed to transform data: {str(e)}", cause=e)
